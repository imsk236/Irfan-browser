"""Tests for export service output (CSV content, JSON structure, Excel sheets, PDF-HTML).

Service functions (export_csv, export_json, export_excel) call get_engine()
directly — the engine fixture already redirects this to the in-memory database,
so the services see the same seeded test data as the rest of the test session.

Scenario: researcher exports at end of a cataloguing session. Archive contains:
  خزانة مسقط (0001) → volume (80 folios) → رسالة في الأصول (copy_year=1200,
  copy_place=نزوى) → person ابن النضر (kunya=أبو عبدالله) as مؤلف → wilaya مسقط.
"""
import csv
import json
import os
import pytest
import openpyxl

from src.services import volumes as vol_svc
from src.services import works as work_svc
from src.services import persons as person_svc
from src.services import relationships as rel_svc
from src.services import annotations as ann_svc
from src.services import export as csv_svc
from src.services import export_excel as excel_svc
from src.services import export_pdf as pdf_svc


@pytest.fixture
def archive(session):
    """Populated archive for export tests."""
    repo = vol_svc.create_repository(session, "0001", "خزانة مسقط")
    vol = vol_svc.create_volume(session, repo.id, folio_count=80)
    work = work_svc.create_work(
        session, vol.id,
        title="رسالة في الأصول",
        copy_year=1200,
        copy_place="نزوى",
    )
    person = person_svc.create_person(
        session, preferred_name="ابن النضر",
        kunya="أبو عبدالله",
    )
    rel_svc.link_person_to_work(session, person.id, work.id, role="مؤلف")
    person_svc.set_wilayas(session, person.id, ["مسقط"])
    return {"repo": repo, "vol": vol, "work": work, "person": person}


# ── CSV ────────────────────────────────────────────────────────────────────────

def test_csv_creates_files(engine, archive, tmp_path):
    files = csv_svc.export_csv(str(tmp_path))
    assert len(files) > 0
    assert all(f.endswith(".csv") for f in files)


def test_csv_volumes_content(engine, archive, tmp_path):
    csv_svc.export_csv(str(tmp_path))
    volumes_csv = tmp_path / "volumes.csv"
    assert volumes_csv.exists()
    with open(volumes_csv, encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    assert any(r["serial"] == archive["vol"].serial for r in rows)


def test_csv_persons_content(engine, archive, tmp_path):
    csv_svc.export_csv(str(tmp_path))
    persons_csv = tmp_path / "persons.csv"
    assert persons_csv.exists()
    with open(persons_csv, encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    assert any(r["preferred_name"] == "ابن النضر" for r in rows)
    # kunya field should be in the row
    assert any(r.get("kunya") == "أبو عبدالله" for r in rows)


def test_csv_relationships_content(engine, archive, tmp_path):
    csv_svc.export_csv(str(tmp_path))
    rel_csv = tmp_path / "person_relationships.csv"
    assert rel_csv.exists()
    with open(rel_csv, encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    assert any(r["role"] == "مؤلف" for r in rows)


def test_csv_utf8_bom_present(engine, archive, tmp_path):
    """CSV files must start with UTF-8 BOM for Excel Arabic compatibility."""
    csv_svc.export_csv(str(tmp_path))
    volumes_csv = tmp_path / "volumes.csv"
    raw = volumes_csv.read_bytes()
    assert raw[:3] == b"\xef\xbb\xbf", "CSV must begin with UTF-8 BOM"


def test_csv_empty_tables_produce_no_file(engine, tmp_path):
    """Tables with zero rows are skipped — no empty CSV created."""
    files = csv_svc.export_csv(str(tmp_path))
    # In a fresh DB (no archive fixture) only vocab is seeded
    filenames = [os.path.basename(f) for f in files]
    # works and volumes tables start empty in a fresh DB
    assert "works.csv" not in filenames
    assert "volumes.csv" not in filenames


def test_csv_trailing_slash_in_path_ok(engine, archive, tmp_path):
    """Trailing slash in output_dir is normalised cleanly."""
    files = csv_svc.export_csv(str(tmp_path) + os.sep)
    assert len(files) > 0
    # No double-separator in any path
    for f in files:
        assert "//" not in f.replace("\\\\", "")


# ── JSON ───────────────────────────────────────────────────────────────────────

def test_json_creates_file(engine, archive, tmp_path):
    out = tmp_path / "archive.json"
    result = csv_svc.export_json(str(out))
    assert os.path.exists(result)
    assert result.endswith(".json")


def test_json_top_level_keys_are_table_names(engine, archive, tmp_path):
    out = tmp_path / "archive.json"
    csv_svc.export_json(str(out))
    with open(out, encoding="utf-8") as f:
        data = json.load(f)
    for key in ("volumes", "works", "persons", "person_relationships"):
        assert key in data, f"Expected table '{key}' in JSON export"


def test_json_volumes_contains_serial(engine, archive, tmp_path):
    out = tmp_path / "archive.json"
    csv_svc.export_json(str(out))
    with open(out, encoding="utf-8") as f:
        data = json.load(f)
    serials = [v["serial"] for v in data["volumes"]]
    assert archive["vol"].serial in serials


def test_json_persons_contains_preferred_name(engine, archive, tmp_path):
    out = tmp_path / "archive.json"
    csv_svc.export_json(str(out))
    with open(out, encoding="utf-8") as f:
        data = json.load(f)
    names = [p["preferred_name"] for p in data["persons"]]
    assert "ابن النضر" in names


def test_json_relationships_non_empty(engine, archive, tmp_path):
    out = tmp_path / "archive.json"
    csv_svc.export_json(str(out))
    with open(out, encoding="utf-8") as f:
        data = json.load(f)
    assert len(data["person_relationships"]) >= 1


def test_json_empty_db_produces_valid_file(engine, tmp_path):
    """Empty archive exports without error; all tables have empty arrays."""
    out = tmp_path / "empty.json"
    csv_svc.export_json(str(out))
    with open(out, encoding="utf-8") as f:
        data = json.load(f)
    assert data["volumes"] == []
    assert data["persons"] == []


# ── Excel ──────────────────────────────────────────────────────────────────────

def test_excel_creates_xlsx_file(engine, archive, tmp_path):
    result = excel_svc.export_excel(str(tmp_path))
    assert os.path.exists(result)
    assert result.endswith(".xlsx")


def test_excel_has_works_sheet(engine, archive, tmp_path):
    result = excel_svc.export_excel(str(tmp_path))
    wb = openpyxl.load_workbook(result)
    assert "المخطوطات" in wb.sheetnames


def test_excel_has_persons_sheet(engine, archive, tmp_path):
    result = excel_svc.export_excel(str(tmp_path))
    wb = openpyxl.load_workbook(result)
    assert "الأشخاص" in wb.sheetnames


def test_excel_works_sheet_contains_serial(engine, archive, tmp_path):
    result = excel_svc.export_excel(str(tmp_path))
    wb = openpyxl.load_workbook(result)
    ws = wb["المخطوطات"]
    all_values = [str(cell.value or "") for row in ws.iter_rows() for cell in row]
    assert archive["vol"].serial in all_values


def test_excel_works_sheet_contains_work_title(engine, archive, tmp_path):
    result = excel_svc.export_excel(str(tmp_path))
    wb = openpyxl.load_workbook(result)
    ws = wb["المخطوطات"]
    all_values = [str(cell.value or "") for row in ws.iter_rows() for cell in row]
    assert "رسالة في الأصول" in all_values


def test_excel_persons_sheet_contains_name(engine, archive, tmp_path):
    result = excel_svc.export_excel(str(tmp_path))
    wb = openpyxl.load_workbook(result)
    ws = wb["الأشخاص"]
    all_values = [str(cell.value or "") for row in ws.iter_rows() for cell in row]
    assert "ابن النضر" in all_values


def test_excel_researcher_name_in_workbook(engine, archive, tmp_path):
    """Researcher name kwarg is written somewhere in the workbook."""
    result = excel_svc.export_excel(str(tmp_path), researcher_name="د. علي الكندي")
    wb = openpyxl.load_workbook(result)
    all_values = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                all_values.append(str(cell.value or ""))
    assert any("علي الكندي" in v for v in all_values)


# ── PDF-HTML ───────────────────────────────────────────────────────────────────

def test_pdf_html_200(client, archive):
    r = client.get("/export/pdf-html")
    assert r.status_code == 200


def test_pdf_html_content_type(client, archive):
    r = client.get("/export/pdf-html")
    assert "text/html" in r.headers.get("content-type", "")


def test_pdf_html_contains_serial(client, archive):
    r = client.get("/export/pdf-html")
    assert archive["vol"].serial in r.text


def test_pdf_html_contains_arabic_text(client, archive):
    r = client.get("/export/pdf-html")
    html_content = r.text
    assert "<html" in html_content.lower()
    # Body must contain Arabic characters
    assert any(ord(c) > 0x600 for c in html_content)


def test_pdf_html_empty_db_returns_200_not_500(client):
    """Empty archive renders PDF HTML without crashing."""
    r = client.get("/export/pdf-html")
    assert r.status_code == 200
    assert "<html" in r.text.lower()


def test_pdf_html_lists_all_scribes_not_just_last(client, session, archive):
    """A work with multiple ناسخ must show every scribe name, not just the last one."""
    scribe1 = person_svc.create_person(session, preferred_name="الناسخ الأول")
    scribe2 = person_svc.create_person(session, preferred_name="الناسخ الثاني")
    rel_svc.link_person_to_work(session, scribe1.id, archive["work"].id, role="ناسخ")
    rel_svc.link_person_to_work(session, scribe2.id, archive["work"].id, role="ناسخ")

    r = client.get("/export/pdf-html")
    assert "الناسخ الأول" in r.text
    assert "الناسخ الثاني" in r.text
