import os
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import text

from ..db.engine import get_engine
from .export import validate_export_path

_DARK_GREEN = "2D3B2D"
_MID_GREEN = "4A5E4A"
_CREAM = "F7F5EF"
_LOGO_PATH = Path(__file__).parent.parent / "assets" / "irfan_logo.png"


def export_excel(output_dir: str, researcher_name: str = "") -> str:
    path = validate_export_path(output_dir)
    path.mkdir(parents=True, exist_ok=True)
    engine = get_engine()

    with engine.connect() as conn:
        works_rows = list(conn.execute(text("""
            SELECT v.serial, r.name AS repo_name, v.repository_volume_number,
                   v.folio_count, w.id AS work_id, w.title,
                   w.topic_category, w.copy_place, w.copy_date_as_written, w.copy_year
            FROM works w
            JOIN volumes v ON v.id = w.volume_id
            JOIN repositories r ON r.id = v.repository_id
            ORDER BY v.serial, w.id
        """)).mappings())

        copyist_rows = list(conn.execute(text("""
            SELECT rel.work_id, p.preferred_name
            FROM person_relationships rel
            JOIN persons p ON p.id = rel.person_id
            WHERE rel.role = 'ناسخ' AND rel.level = 'work'
        """)).mappings())
        copyist_map = {r["work_id"]: r["preferred_name"] for r in copyist_rows}

        person_rows = list(conn.execute(text("""
            SELECT p.preferred_name, p.kunya, p.laqab,
                   p.birth_year_earliest, p.death_year_earliest,
                   GROUP_CONCAT(pw.wilaya, '، ') AS wilayas,
                   COUNT(DISTINCT rel.id) AS appearances
            FROM persons p
            LEFT JOIN person_wilayas pw ON pw.person_id = p.id
            LEFT JOIN person_relationships rel ON rel.person_id = p.id
            GROUP BY p.id
            ORDER BY p.preferred_name
        """)).mappings())

    wb = openpyxl.Workbook()
    _build_works_sheet(wb.active, works_rows, copyist_map, researcher_name)
    _build_persons_sheet(wb.create_sheet("الأشخاص"), person_rows)

    export_date = datetime.now().strftime("%Y-%m-%d_%H-%M")
    out_path = os.path.join(str(path), f"ارشيف_عرفان_{export_date}.xlsx")
    wb.save(out_path)
    return out_path


def _add_logo(ws, col_count: int) -> None:
    if not _LOGO_PATH.exists():
        return
    img = XLImage(str(_LOGO_PATH))
    img.width = 48
    img.height = 48
    ws.row_dimensions[1].height = 38
    ws.add_image(img, "A1")


def _title_row(ws, col_count: int, text_val: str, row: int = 1) -> None:
    col_letter = get_column_letter(col_count)
    ws.merge_cells(f"A{row}:{col_letter}{row}")
    cell = ws[f"A{row}"]
    cell.value = text_val
    cell.font = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    cell.fill = PatternFill("solid", fgColor=_DARK_GREEN)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28


def _header_row(ws, headers: list[str], widths: list[int], row: int = 2) -> None:
    for col_idx, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        cell.fill = PatternFill("solid", fgColor=_MID_GREEN)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[row].height = 20


def _build_works_sheet(ws, works_rows, copyist_map, researcher_name: str) -> None:
    ws.title = "المخطوطات"
    ws.sheet_view.rightToLeft = True

    _add_logo(ws, 9)

    export_date = datetime.now().strftime("%Y-%m-%d")
    title = "أرشيف عرفان — فهرس المخطوطات"
    if researcher_name:
        title += f"  |  {researcher_name}"
    title += f"  |  {export_date}"

    _title_row(ws, 9, title, row=2)

    headers = [
        "الرقم التسلسلي", "الخزانة", "رقم المجلد", "عدد الأوراق",
        "عنوان المصنَّف", "التصنيف الموضوعي", "مكان النسخ", "تاريخ النسخ", "الناسخ",
    ]
    widths = [18, 20, 14, 12, 38, 22, 18, 18, 22]
    _header_row(ws, headers, widths, row=3)

    for i, row in enumerate(works_rows):
        r = i + 4
        copy_date = row["copy_date_as_written"] or (str(row["copy_year"]) if row["copy_year"] else "")
        values = [
            row["serial"],
            row["repo_name"],
            row["repository_volume_number"],
            row["folio_count"],
            row["title"],
            row["topic_category"] or "",
            row["copy_place"] or "",
            copy_date,
            copyist_map.get(row["work_id"], ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="top")
            if i % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=_CREAM)

    ws.freeze_panes = "A4"


def _build_persons_sheet(ws, person_rows) -> None:
    ws.title = "الأشخاص"
    ws.sheet_view.rightToLeft = True

    _add_logo(ws, 7)
    _title_row(ws, 7, "أرشيف عرفان — فهرس الأشخاص", row=2)

    headers = ["الاسم المعتمد", "الكنية", "اللقب", "سنة الميلاد", "سنة الوفاة", "الولايات", "عدد الظهورات"]
    widths = [30, 18, 18, 14, 14, 28, 14]
    _header_row(ws, headers, widths, row=3)

    for i, row in enumerate(person_rows):
        r = i + 4
        values = [
            row["preferred_name"],
            row["kunya"] or "",
            row["laqab"] or "",
            row["birth_year_earliest"] or "",
            row["death_year_earliest"] or "",
            row["wilayas"] or "",
            row["appearances"] or 0,
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="top")
            if i % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=_CREAM)

    ws.freeze_panes = "A4"
